#!/usr/bin/python3

import pandas as pd
import openpyxl as opx
import numpy as np

def get_data(wb,name):
    xsheet = wb['x-axis']
    for x in xsheet.values:
        if x[0] is not None:
            xaxis = np.array(x)
            break
    xlen = len(xaxis)
    ysheet = wb['y-axis']
    for y in ysheet.values:
        if y[0] is not None:
            yaxis = np.array(y)
            break
    ylen = len(yaxis)

    costOptSheet = wb['cost($)-opt']
    costOpt = np.zeros((xlen,ylen))
    for j,co in enumerate(costOptSheet.values):
        c = np.array(co).astype(float)
        #c[np.isnan(c)] = -100.0
        c[c < 1e-6] = np.nan
        costOpt[:,j] = c
    
    nstagesOptSheet = wb['nStages(-)-opt']
    nstagesOpt = np.zeros((xlen,ylen))
    for j,co in enumerate(nstagesOptSheet.values):
        c = np.array(co).astype(int)
        c[c == np.array(None)] = -10
        c[c == 0] = -10
        nstagesOpt[:,j] = c
    
    ccrOptSheet = wb['CCR(-)-opt']
    ccrOpt = np.zeros((xlen,ylen))
    for j,co in enumerate(ccrOptSheet.values):
        c = np.array(co).astype(float)
        #c[np.isnan(c)] = -100.0
        c[c < 1e-6] = np.nan
        ccrOpt[:,j] = c
    
    cost90Sheet = wb['cost($)-CCR90']
    cost90 = np.zeros((xlen,ylen))
    for j,co in enumerate(cost90Sheet.values):
        c = np.array(co).astype(float)
        #c[np.isnan(c)] = -100.0
        c[c < 1e-6] = np.nan
        cost90[:,j] = c
    
    nstages90Sheet = wb['nStages(-)-CCR90']
    nstages90 = np.zeros((xlen,ylen))
    for j,co in enumerate(nstages90Sheet.values):
        c = np.array(co).astype(int)
        c[c == np.array(None)] = -10
        c[c == 0] = -10
        nstages90[:,j] = c
    
    costRedOpt90Sheet = wb['costReductionVs90pct(-)-opt']
    costRedOpt90 = np.zeros((xlen,ylen))
    for j,co in enumerate(costRedOpt90Sheet.values):
        c = np.array(co).astype(float)
        #c[np.isnan(c)] = -100.0
        c[c > -1e-30] = np.nan
        costRedOpt90[:,j] = c

    # For some reason, Excel sheet y and x is transposed
    data_dict = {'xval':yaxis,'yval':xaxis,'xlen':ylen,'ylen':xlen,
            'costOpt':costOpt,'nstagesOpt':nstagesOpt,'ccrOpt':ccrOpt,
            'cost90':cost90,'nstages90':nstages90,
            'costRedOpt90':costRedOpt90,'appl':name}

    return data_dict

if __name__ == "__main__":
    all_data = {}
    cembook = opx.load_workbook('allCementResults.xlsx')
    cem = get_data(cembook,'cem')
    all_data['cem'] = cem
    steelbook = opx.load_workbook('allSteelResults.xlsx')
    steel = get_data(steelbook,'steel')
    all_data['steel'] = steel
    coalbook = opx.load_workbook('allCoalResults.xlsx')
    coal = get_data(coalbook,'coal')
    all_data['coal'] = coal
    fccbook = opx.load_workbook('allFCCResults.xlsx')
    fcc = get_data(fccbook,'fcc')
    all_data['fcc'] = fcc
    fgbook = opx.load_workbook('allFGResults.xlsx')
    fg = get_data(fgbook,'fg')
    all_data['fg'] = fg
    lsfobook = opx.load_workbook('allLSFOResults.xlsx')
    lsfo = get_data(lsfobook,'lsfo')
    all_data['lsfo'] = lsfo

    np.save('all-data-dict.npy',all_data)
